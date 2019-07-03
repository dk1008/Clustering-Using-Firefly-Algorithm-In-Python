

				#################################################################################
				# 										#
				#		Clustering Using Firefly Algorithm In Python			#
				#										#
				#################################################################################


#*******************************************************************************************************************************************#
'''***************************************************************************************************************************************'''
#*******************************************************************************************************************************************#

import matplotlib.pyplot as plt
import openpyxl
import random
import math
print(" ")
print("Example of Path of Excel File: /home/lenovo/Documents/K_Mean_Firefly_clustering/glass_Data_set.xlsx")
Location = input("Enter Path Of Excel File: ")
wb = openpyxl.load_workbook(Location)
sheet = wb.active





###################
## It Contains Data Set Values
NData = sheet.max_row
###################





###################
## It Take Dimension Of Data Set As Input
DD = sheet.max_column
###################





###################
## It Take Number Of Firefly as Input 
MaxGeneration = int(input("Enter MaxGeneration Value: "))
###################





###################
## It Take Number Of Firefly as Input 
NF = int(input("Enter Number Of Firefly: "))
###################





###################
## It Take Number Of Centroid As Input
NC = int(input("Every Firefly Has Number Of Centriod: "))
###################





###################
#
ALPHA = .7 #randomization parameter
BETA = .5 #attractiveness
GAMMA = 2 #light absorption cofficient
###################





#START****************************************************************************************************************************************
'''Here UB and LB contains Upper and lower bound of respective columns.'''
# LB -> Lower Bound 	UB -> Upper Bound
LB = []	# It Take Lower Bound As Input
UB = []	# It take Upper Bound As Input
for i in range(DD):
	for j in range(NData):
		if(j == 0):
			LB.append(sheet.cell(row = j+1, column = i+1).value)
			UB.append(sheet.cell(row = j+1, column = i+1).value)
		else:
			num = sheet.cell(row = j+1, column = i+1).value
			LB[i] = min(LB[i] , num)
			UB[i] = max(UB[i] , num)

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************

MD = 0	#MD -> Max Distance Of Centroid DataSets 

''' This Will Generate A Distance Which Will Be longest Distance Between any Points in centroid Data set'''
for i in range(DD):
	MD += math.pow((UB[i]-LB[i]),2)
MD = math.sqrt(MD)

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
'''Distance function will calculate Between two Point or two array of same length.'''
def Distance_Evaluation(CP , DP):	#CP -> Centroid Data Set	#DP -> Data Point DataSet
	sum = 0
	for i in range(DD):
		sum += math.pow((CP[i]-DP[i]),2)
	return (math.sqrt(abs(sum))) 	#return square of distance between two firefly..

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************

DPP = []	#DPP -> Input Data

#Generating Data Set
#DPP Contain "NData" Unit Of Data Set Value
#Each Data Set Has "DD" Unit Dimension
def Data_Point_Position(Location, wb, sheet, NData , DD):
	for i in range (NData):
		DPP.append([])
		for j in range (DD):
			DPP[i].append([])
			DPP[i][j] = sheet.cell(row = i+1, column = j+1).value #initializing the Input Data

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

LOC = []	#LOC -> List Of Centroid

#generating Fireflies
#LOC List Contain "NF" Unit of Firefly
#Each Firefly has "NC" Unit of Centroid
#Each Centroid has "DD" Unit of Dimension
def Number_Of_Firefly(NF , NC , DD , LB , UB):
	for i in range(NF):
		LOC.append([])
		for j in range(NC):
			LOC[i].append([])
			for k in range(DD):
				LOC[i][j].append([])
				LOC[i][j][k] = random.uniform(LB[k] , UB[k])

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

CRD = []	#CDL -> Centroid With Respective DataSet
def respective_Centroid_List(NC):
	CRD [:]=[]
	for i in range(NC):
		CRD.append([])

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

DBCD = []	#DBCD -> Distance Between Centroid And Data Points
def Respective_Centroid_Distance(NF ):
	DBCD[:]=[]
	for i in range(NF):
		DBCD.append([])
		DBCD[i] = 0

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

#Seprating Dataset According to Their Respective Centroid
#MD -> Max Distance Of Centroid Data Sets
#NData -> Number Of DataSet
#NC -> Each Firefly Has Number of centroid 
def DataSet_Centroid_Connection(MD , NData , NC ):
	for i in range(NData):
		CT_Index = 0	#Allocate A DataSet To A Respective Centroid
		dis = MD
		for j in range(NC):
			distance = Distance_Evaluation(DPP[i] , Centroid[j])
			if(distance <= dis):
				dis = distance
				CT_Index = j
		CRD[CT_Index].append(DPP[i])
	 
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

#Calculating and Allocating Dataset-Centroid Distance According to Their Respective Centroid
#MD -> Max Distance Of Centroid Data Sets
#NF -> Number Of Firefly
#NData -> Number Of DataSet
#NC -> Each Firefly Has Number of centroid 
def DataSet_Centroid_Distance_Connection(MD , NF , NData , NC ):
	for i in range(NF):
		for j in range(NData):
			dis = MD
			for k in range(NC):
				distance = Distance_Evaluation(DPP[j] , LOC[i][k])
				if(distance <= dis):
					dis = distance
			DBCD[i] += dis 

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
Error_Value = []	#SSE -> sum of square root of distance between centroid and input data.
#END-----------------------------------------------------------------------------------------------------------------------------------------





#START***************************************************************************************************************************************
'''function will replace old centroid by new centroid by calculating average of data in respective centroid.'''
def Centroid_Average(NC , DD):
	for i in range(NC):
		for j in range(DD):
			total = 0
			avg = 0
			for k in range(len(CRD[i])):
				total += CRD[i][k][j]
			if(total != 0):
				avg = total/len(CRD[i])
			Centroid[i][j] = avg

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
def Display(NC , DD):
	print(" ")
	for i in range(NC):
		print("Centroid-{0} is: ".format(i+1),end = " "),
		for j in range(DD):
			print("%.6f  "%(Centroid[i][j]), end = " "),
		print(" ")
			
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
def Cluster_Ploting(NC): #here we have to plot all input data and centroid in different color..
	#plt.axis([math.floor(LB[0]),math.ceil(UB[0]),math.floor(LB[1]),math.ceil(UB[1])])
	for i in range(NData):
		plt.plot(DPP[i][0],DPP[i][1],"r o")

	for i in range(NC):
		for j in range(len(CRD[i])):
			if(len(CRD[i])>0):
				x = [Centroid[i][0],CRD[i][j][0]]
				y = [Centroid[i][1],CRD[i][j][1]]
				plt.plot(x,y,"y-")
	for i in range(NC):
		plt.plot(Centroid[i][0] , Centroid[i][1],"g*")
	plt.show()
		
'''END----------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
#Siloutee Value plotting...
def SI_Ploting():
	for i in range(len(SI)):
		if(i < MaxGeneration):
			plt.plot(i,SI[i],"g *")
		else:
			plt.plot(i,SI[i],"r o")
	plt.show()
'''END----------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
def Error_Ploting(MaxGeneration):
	for i in range(len(Error_Value)):
		if(i<= MaxGeneration):
			plt.plot(i,Error_Value[i],"g *")
		else:
			plt.plot(i,Error_Value[i],"r o")
	plt.show()
'''END----------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
test = []
def Test_Centroid(NC , DD):
	for i in range(NC):
		test.append([])
		for j in range(DD):
			test[i].append(0)

'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START*************************************************************************************************************************************
IntraCD = []
for i in range(NC):
	IntraCD.append([])
def Intra_Cluster_distance(NC):
	for i in range(NC):
		for j in range(len(CRD[i])):
			total = 0
			if(len(CRD[i]) > 0):
				for k in range(len(CRD[i])):
					dis = Distance_Evaluation(CRD[i][j] , CRD[i][k])
					total += dis
			IntraCD[i].append(total)
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START**********************************************************************************************************************************
InterCD = []
for i in range(NC):
	InterCD.append([])
def Inter_Cluster_distance(NC):
	for i in range(NC):
		for j in range(len(CRD[i])):
			total = 0
			if(len(CRD[i]) > 0):
				for k in range(NC):
					if(i != k):
						for l in range(len(CRD[k])):
							if(len(CRD[k]) > 0):
								dis = Distance_Evaluation(CRD[i][j] , CRD[k][l])
								total += dis
				InterCD[i].append(total)
'''END---------------------------------------------------------------------------------------------------------------------------------'''





#START***********************************************************************************************************************************
SI = []
def siloutte_Indexing(NC):
	total = 0
	x = 0
	Intra_Cluster_distance(NC)
	Inter_Cluster_distance(NC)
	for i in range(NC):
		for j in range(len(IntraCD[i])):
			if(len(IntraCD[i]) > 0):
				x = (InterCD[i][j] - IntraCD[i][j])/max(IntraCD[i][j],InterCD[i][j])
		total += x
		total = total/NC
	SI.append(total) 
'''END---------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************

Data_Point_Position(Location, wb, sheet, NData , DD)
Number_Of_Firefly(NF , NC , DD , LB , UB)
Respective_Centroid_Distance(NF )
DataSet_Centroid_Distance_Connection(MD , NF , NData , NC )
Error_Value.append(min(DBCD))
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************
def saving_Error_in_excel():
	result_location = ("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_FE_K.xlsx")
	wx = openpyxl.load_workbook(result_location)
	sheet1 = wx.active
	no_of_row = sheet1.max_row
	c1 = sheet1.cell(row = no_of_row+1, column = 2)
	c1.value = Error_Value[len(Error_Value)-1]
	wx.save("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_FE_K.xlsx")
'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START****************************************************************************************************************************************
def saving_file_in_excel():
	result_location = ("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_F_K.xlsx")
	ws = openpyxl.load_workbook(result_location)
	sheet1 = ws.active
	no_of_row = sheet1.max_row
	c1 = sheet1.cell(row = no_of_row+1, column = 2)
	c1.value = SI[len(SI)-1]
	ws.save("/home/lenovo/Documents/K_Mean_Firefly_clustering/RESULT_F_K.xlsx")
'''END------------------------------------------------------------------------------------------------------------------------------------'''




Centroid = None
#START***************************************************************************************************************************************
#Firefly Algorithm.....
for i in range(MaxGeneration):
	for j in range(NF):
		for k in range(NF):
			if(DBCD[j] > DBCD[k]):
				for l in range(NC):
					for m in range(DD):
						r_square = (Distance_Evaluation(LOC[j][l] , LOC[k][l]))**2
						LOC[j][l][m] += BETA*math.exp(-GAMMA*r_square)*(LOC[k][l][m]-LOC[j][l][m]) + BETA*math.exp(-GAMMA*r_square)*(LOC[k][l][m]-LOC[j][l][m]) + ALPHA*(random.uniform(0,1) - .5)
						if(LOC[j][l][m] < LB[m]):
							LOC[j][l][m] = random.uniform(LB[m],UB[m])
						if(LOC[j][l][m] > UB[m]):
							LOC[j][l][m] = random.uniform(LB[m],UB[m])
	Respective_Centroid_Distance(NF )
	DataSet_Centroid_Distance_Connection(MD , NF , NData , NC )
	Error_Value.append(min(DBCD))	#Error value calculating.
	
	Centroid = LOC[DBCD.index(min(DBCD))]
	respective_Centroid_List(NC)
	DataSet_Centroid_Connection(MD , NData , NC )
	siloutte_Indexing(NC)  #Siloutee value calculating.


'''END------------------------------------------------------------------------------------------------------------------------------------'''





#START**************************************************************************************************************************************
def K_Means_DataSet_Centroid_Distance_Connection(MD , NData , NC ):
	total = 0
	for i in range(NData):
		dis = MD
		for j in range(NC):
			distance = Distance_Evaluation(DPP[i] , Centroid[j])
			if(distance <= dis):
				dis = distance
		total += dis
	return total
'''END-----------------------------------------------------------------------------------------------------------------------------------'''





#START***************************************************************************************************************************************
#K-Mean Algorithm.....
Centroid = LOC[DBCD.index(min(DBCD))]
Test_Centroid(NC , DD)
Display(NC , DD)
loop = 0
while(True):
	respective_Centroid_List(NC)
	DataSet_Centroid_Connection(MD , NData , NC )
	Centroid_Average(NC , DD)
	k_dis = K_Means_DataSet_Centroid_Distance_Connection(MD , NData , NC )
	Error_Value.append(k_dis)
	siloutte_Indexing(NC)  #Siloutee value calculating.
	Display(NC, DD)
	if(DD == 2):
		Cluster_Ploting(NC)
	flag = 0
	for i in range(NC):
		for j in range(DD):
			if(Centroid[i][j] != test[i][j]):
				test[i][j] = Centroid[i][j]
				flag = 1
	if(flag == 0):
		if(loop < 5):
			loop += 1
		else:
			break	


print(" ")
saving_file_in_excel()
saving_Error_in_excel()
Error_Ploting(MaxGeneration)	#Error Value Displaying.
SI_Ploting()  # Siloutee value displaying.

'''
print("MAX DISTANCE ", MD)
print()
print("Input Data is")
print(DPP)
print()
print("Centroid with respective data")
print(CRD)
print()
print("DIstance between centroid and input data")
print(DBCD)'''
'''END------------------------------------------------------------------------------------------------------------------------------------'''
